"""
app/services/engines/media_convert.py

Real image, video, audio, and light PDF-conversion handling. Replaces the
earlier AVX stub: AVX was never confirmed as a real maintained project
(see docs/engines.md), so this uses a verified, properly-licensed stack
instead — Pillow (HPND/permissive) for images, ffmpeg (subprocess only;
see docs/licensing.md for the GPL-via-subprocess note) for video/audio,
poppler-utils' pdftoppm (subprocess; GPL-2.0) for PDF rasterization,
pypdf/reportlab for PDF text, openpyxl for spreadsheets, rembg + onnxruntime
(both MIT) for image_bg_remove — see that handler's docstring for the
model-licensing check.

One tool is left as a documented stub rather than faked: audio_to_text
(needs a speech-to-text model, e.g. Whisper — the standard hosts for
Whisper/faster-whisper's model weights, huggingface.co and
openaipublic.azureedge.net, are both unreachable from this sandbox's
network egress allowlist, so unlike image_bg_remove below this genuinely
couldn't be verified working here; see docs/engines.md).
"""
from __future__ import annotations

import csv
import difflib
import io
import zipfile
from functools import lru_cache
from typing import Any, BinaryIO

from PIL import Image, ImageDraw, ImageFont, UnidentifiedImageError
from pypdf import PdfReader
from pypdf.errors import PyPdfError
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from services.tool_timeouts import get_subprocess_timeout_seconds

from ._util import find_ttf_font, run, scratch_dir
from .base import Engine, EngineResult

# u2netp ("portrait", ~4.5MB) rather than rembg's default u2net (~176MB):
# image_bg_remove is a synchronous tool (image-category, not in
# ASYNC_TOOL_NAMES — see routes/tools.py) that resolves inline in the HTTP
# request, so bounded latency matters more here than the small quality
# edge u2net has on complex scenes. Verified against this sandbox's
# network egress: the model downloads from rembg's GitHub release
# (github.com/danielgatis/rembg/releases), which is reachable, unlike
# audio_to_text's Whisper hosts (see the module docstring).
REMBG_MODEL_NAME = "u2netp"


@lru_cache
def _rembg_session():
    """Cached across requests in the same process — rembg's own on-disk
    cache (~/.rembg/models) means the .onnx file itself is only ever
    downloaded once regardless, but constructing the onnxruntime
    InferenceSession has its own non-trivial cost that's worth not paying
    on every single request."""
    from rembg import new_session

    return new_session(REMBG_MODEL_NAME)

IMAGE_FORMAT_MIME = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "webp": "image/webp", "gif": "image/gif", "bmp": "image/bmp", "tiff": "image/tiff",
    # Both built into this Pillow build (12.3.0) with no extra plugin —
    # verified empirically (ICO is core Pillow; AVIF ships via bundled
    # libavif in this wheel, confirmed with PIL.features.check("avif")
    # before relying on it) rather than assumed from the Pillow version
    # number alone.
    "ico": "image/x-icon", "avif": "image/avif",
}


def _pil_format(fmt: str) -> str:
    return {"jpg": "JPEG"}.get(fmt.lower(), fmt.upper())


class MediaConvertEngine(Engine):
    name = "media_convert"

    def __init__(self) -> None:
        self._handlers = {
            # pdf <-> image / text
            "pdf_to_jpg": self._pdf_to_image,
            "pdf_to_png": self._pdf_to_image,
            "pdf_to_tiff": self._pdf_to_image,
            "jpg_to_pdf": self._image_to_pdf,
            "png_to_pdf": self._image_to_pdf,
            "webp_to_pdf": self._image_to_pdf,
            "gif_to_pdf": self._image_to_pdf,
            "bmp_to_pdf": self._image_to_pdf,
            "tiff_to_pdf": self._image_to_pdf,
            "image_to_pdf": self._image_to_pdf,
            "pdf_to_text": self._pdf_to_text,
            "text_to_pdf": self._text_to_pdf,
            "pdf_compare": self._pdf_compare,
            # image
            "image_convert": self._image_convert,
            "webp_convert": self._image_convert,
            # Specific named format pairs — same handler as image_convert,
            # engine_op supplies the fixed target_format (see
            # tool_router.py's engine_op="key=value" auto-population and
            # tools_catalog.py's module docstring for why these aren't
            # just aliases).
            "png_to_jpg": self._image_convert,
            "jpg_to_png": self._image_convert,
            "png_to_webp": self._image_convert,
            "webp_to_png": self._image_convert,
            "bmp_to_png": self._image_convert,
            "png_to_bmp": self._image_convert,
            "tiff_to_png": self._image_convert,
            "png_to_tiff": self._image_convert,
            "gif_to_png": self._image_convert,
            # Remaining pairs among the same six formats — filling out the
            # ordered-pair grid, same handler.
            "jpg_to_webp": self._image_convert,
            "jpg_to_gif": self._image_convert,
            "jpg_to_bmp": self._image_convert,
            "jpg_to_tiff": self._image_convert,
            "png_to_gif": self._image_convert,
            "webp_to_jpg": self._image_convert,
            "webp_to_gif": self._image_convert,
            "webp_to_bmp": self._image_convert,
            "webp_to_tiff": self._image_convert,
            "gif_to_jpg": self._image_convert,
            "gif_to_webp": self._image_convert,
            "gif_to_bmp": self._image_convert,
            "gif_to_tiff": self._image_convert,
            "bmp_to_jpg": self._image_convert,
            "bmp_to_webp": self._image_convert,
            "bmp_to_gif": self._image_convert,
            "bmp_to_tiff": self._image_convert,
            "tiff_to_jpg": self._image_convert,
            "tiff_to_webp": self._image_convert,
            "tiff_to_gif": self._image_convert,
            "tiff_to_bmp": self._image_convert,
            # New formats this pass: ICO (favicons) and AVIF (modern web
            # images) — both confirmed working in this Pillow build with
            # no new dependency, see IMAGE_FORMAT_MIME's comment above.
            "png_to_ico": self._image_convert,
            "ico_to_png": self._image_convert,
            "jpg_to_ico": self._image_convert,
            "png_to_avif": self._image_convert,
            "avif_to_png": self._image_convert,
            "jpg_to_avif": self._image_convert,
            "avif_to_jpg": self._image_convert,
            "webp_to_avif": self._image_convert,
            "avif_to_webp": self._image_convert,
            "image_resize": self._image_resize,
            "image_compress": self._image_compress,
            "image_crop": self._image_crop,
            "image_watermark": self._image_watermark,
            "image_rotate": self._image_rotate,
            "heic_to_jpg": self._heic_to_jpg,
            "svg_to_png": self._svg_to_png,
            "image_bg_remove": self._image_bg_remove,
            # video
            "video_compress": self._video_compress,
            "video_convert": self._video_convert,
            "video_trim": self._video_trim,
            "video_to_gif": self._video_to_gif,
            "video_extract_audio": self._video_extract_audio,
            "video_merge": self._video_merge,
            "video_resize": self._video_resize,
            "video_watermark": self._video_watermark,
            "video_mute": self._video_mute,
            "subtitle_burn": self._subtitle_burn,
            # Specific named video format pairs — same handler as
            # video_convert, engine_op supplies target_format. ffmpeg
            # sniffs real content regardless of the scratch file's
            # extension (verified — see test_engines.py), so the
            # source-side name in each pair below is documentation, not
            # something the handler itself needs to know.
            "mp4_to_webm": self._video_convert,
            "webm_to_mp4": self._video_convert,
            "mp4_to_mkv": self._video_convert,
            "mkv_to_mp4": self._video_convert,
            "mp4_to_mov": self._video_convert,
            "mov_to_mp4": self._video_convert,
            # Remaining pairs among mp4/webm/mkv/mov, plus two more
            # containers (avi, flv) bidirectional with mp4 — same handler.
            "webm_to_mkv": self._video_convert,
            "mkv_to_webm": self._video_convert,
            "webm_to_mov": self._video_convert,
            "mov_to_webm": self._video_convert,
            "mkv_to_mov": self._video_convert,
            "mov_to_mkv": self._video_convert,
            "mp4_to_avi": self._video_convert,
            "avi_to_mp4": self._video_convert,
            "mp4_to_flv": self._video_convert,
            "flv_to_mp4": self._video_convert,
            # New containers this pass — plain ffmpeg -i in -> out with no
            # extra args (same as every pair above); 3gp was tried and
            # dropped (needs explicit -c:v/-c:a codec args the generic
            # handler doesn't pass, so it's not a same-handler fit — see
            # docs/TODO.md).
            "mp4_to_wmv": self._video_convert,
            "wmv_to_mp4": self._video_convert,
            "mp4_to_ts": self._video_convert,
            "ts_to_mp4": self._video_convert,
            "mp4_to_m4v": self._video_convert,
            "m4v_to_mp4": self._video_convert,
            # Named "video to mp3" extraction pairs — huge real-world
            # search volume ("convert mp4 to mp3") that the generic
            # video_extract_audio tool doesn't capture on its own. Same
            # handler, source container doesn't matter (ffmpeg content-
            # sniffs, same reasoning as every pair above).
            "mp4_to_mp3": self._video_extract_audio,
            "mov_to_mp3": self._video_extract_audio,
            "webm_to_mp3": self._video_extract_audio,
            "mkv_to_mp3": self._video_extract_audio,
            "avi_to_mp3": self._video_extract_audio,
            # Remaining containers — completes "extract audio to mp3" for
            # every video container this catalog supports.
            "flv_to_mp3": self._video_extract_audio,
            "wmv_to_mp3": self._video_extract_audio,
            "ts_to_mp3": self._video_extract_audio,
            "m4v_to_mp3": self._video_extract_audio,
            # audio
            "audio_convert": self._audio_convert,
            "audio_compress": self._audio_compress,
            "audio_trim": self._audio_trim,
            "audio_merge": self._audio_merge,
            "audio_normalize": self._audio_normalize,
            "audio_to_text": self._not_implemented(
                "Needs a speech-to-text model (e.g. Whisper) — out of scope for this pass"
            ),
            # Specific named audio format pairs — same handler as
            # audio_convert.
            "mp3_to_wav": self._audio_convert,
            "wav_to_mp3": self._audio_convert,
            "wav_to_flac": self._audio_convert,
            "flac_to_mp3": self._audio_convert,
            "wav_to_ogg": self._audio_convert,
            "ogg_to_mp3": self._audio_convert,
            "wav_to_m4a": self._audio_convert,
            "m4a_to_mp3": self._audio_convert,
            # Remaining pairs among mp3/wav/flac/ogg/m4a — same handler.
            "mp3_to_flac": self._audio_convert,
            "flac_to_wav": self._audio_convert,
            "mp3_to_ogg": self._audio_convert,
            "ogg_to_wav": self._audio_convert,
            "mp3_to_m4a": self._audio_convert,
            "m4a_to_wav": self._audio_convert,
            "flac_to_ogg": self._audio_convert,
            "ogg_to_flac": self._audio_convert,
            "flac_to_m4a": self._audio_convert,
            "m4a_to_flac": self._audio_convert,
            "ogg_to_m4a": self._audio_convert,
            "m4a_to_ogg": self._audio_convert,
            # New formats this pass — same handler.
            "mp3_to_opus": self._audio_convert,
            "opus_to_mp3": self._audio_convert,
            "mp3_to_aac": self._audio_convert,
            "aac_to_mp3": self._audio_convert,
            "mp3_to_wma": self._audio_convert,
            "wma_to_mp3": self._audio_convert,
            "mp3_to_aiff": self._audio_convert,
            "aiff_to_mp3": self._audio_convert,
            # Connects opus/aac/wma/aiff to the wav hub too, not just mp3
            # (flac/ogg/m4a already had both — this pass closes the gap).
            "wav_to_opus": self._audio_convert,
            "opus_to_wav": self._audio_convert,
            "wav_to_aac": self._audio_convert,
            "aac_to_wav": self._audio_convert,
            "wav_to_wma": self._audio_convert,
            "wma_to_wav": self._audio_convert,
            "wav_to_aiff": self._audio_convert,
            "aiff_to_wav": self._audio_convert,
            # spreadsheets
            "csv_to_xlsx": self._csv_to_xlsx,
            "xlsx_to_csv": self._xlsx_to_csv,
        }

    def process(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        tool_name = options.get("tool_name")
        handler = self._handlers.get(tool_name)
        if handler is None:
            return EngineResult(ok=False, error=f"MediaConvertEngine has no handler for '{tool_name}'")
        try:
            return handler(input_data, options)
        except (PyPdfError, UnidentifiedImageError) as exc:
            # pypdf/Pillow couldn't parse the uploaded file at all — a
            # malformed/corrupted PDF or image, not a bug on our side.
            return EngineResult(ok=False, error=f"{tool_name} failed: {exc}", refundable=False)
        except Exception as exc:  # noqa: BLE001
            return EngineResult(ok=False, error=f"{tool_name} failed: {exc}")

    def _not_implemented(self, reason: str):
        def handler(_input_data, _options):
            return EngineResult(ok=False, error=f"Not implemented: {reason}")

        return handler

    # -- pdf <-> image / text ---------------------------------------------

    # pdftoppm's own natively-supported output formats — not a new
    # capability, just exposing the flags poppler already has (verified
    # with `pdftoppm -h`) instead of only the two this handler used to
    # pick from.
    _PDFTOPPM_FLAGS = {"jpg": "-jpeg", "jpeg": "-jpeg", "png": "-png", "tiff": "-tiff"}

    def _pdf_to_image(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        fmt = options.get("target_format", "jpg").lower()
        poppler_flag = self._PDFTOPPM_FLAGS.get(fmt, "-png")
        with scratch_dir() as d:
            src = d / "in.pdf"
            src.write_bytes(input_data.read())
            run(["pdftoppm", poppler_flag, "-r", "150", str(src), str(d / "page")])

            pages = sorted(d.glob("page-*"))
            if not pages:
                return EngineResult(
                    ok=False, error="No pages rasterized — is this a valid PDF?", refundable=False,
                )

            if len(pages) == 1:
                return EngineResult(
                    ok=True, output_bytes=pages[0].read_bytes(),
                    content_type=IMAGE_FORMAT_MIME.get(fmt, "image/jpeg"),
                )

            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for p in pages:
                    zf.writestr(p.name, p.read_bytes())
            return EngineResult(ok=True, output_bytes=zip_buf.getvalue(), content_type="application/zip",
                                 meta={"pages": len(pages)})

    def _image_to_pdf(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        extra_files: list[bytes] = options.get("extra_files") or []
        images = [Image.open(io.BytesIO(b)).convert("RGB") for b in [input_data.read(), *extra_files]]

        buf = io.BytesIO()
        images[0].save(buf, format="PDF", save_all=True, append_images=images[1:])
        return EngineResult(ok=True, output_bytes=buf.getvalue(), content_type="application/pdf",
                             meta={"pages": len(images)})

    def _pdf_to_text(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        reader = PdfReader(input_data)
        text = "\n\n".join(page.extract_text() or "" for page in reader.pages)
        return EngineResult(ok=True, output_bytes=text.encode("utf-8"), content_type="text/plain",
                             meta={"pages": len(reader.pages)})

    def _text_to_pdf(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        import textwrap

        text = input_data.read().decode("utf-8", errors="replace")
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=letter)
        width, height = letter
        c.setFont("Helvetica", 11)
        y = height - 72
        for paragraph in text.split("\n"):
            for line in textwrap.wrap(paragraph, 95) or [""]:
                if y < 72:
                    c.showPage()
                    c.setFont("Helvetica", 11)
                    y = height - 72
                c.drawString(72, y, line)
                y -= 14
        c.save()
        return EngineResult(ok=True, output_bytes=buf.getvalue(), content_type="application/pdf")

    def _pdf_compare(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        extra_files: list[bytes] = options.get("extra_files") or []
        if not extra_files:
            return EngineResult(
                ok=False, error="pdf_compare needs a second file in options['extra_files']",
                refundable=False,
            )

        a_reader = PdfReader(input_data)
        b_reader = PdfReader(io.BytesIO(extra_files[0]))
        a_text = "\n".join((p.extract_text() or "") for p in a_reader.pages).splitlines()
        b_text = "\n".join((p.extract_text() or "") for p in b_reader.pages).splitlines()

        diff = list(difflib.unified_diff(a_text, b_text, fromfile="file_a.pdf", tofile="file_b.pdf", lineterm=""))
        return EngineResult(
            ok=True, output_bytes="\n".join(diff).encode("utf-8"), content_type="text/plain",
            meta={"changed_lines": sum(1 for line in diff if line.startswith(("+", "-")) and not line.startswith(("+++", "---")))},
        )

    # -- image --------------------------------------------------------------

    def _image_convert(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        target_format = options.get("target_format")
        if not target_format:
            return EngineResult(ok=False, error="options['target_format'] is required", refundable=False)

        img = Image.open(input_data)
        if target_format.lower() in ("jpg", "jpeg") and img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format=_pil_format(target_format))
        return EngineResult(ok=True, output_bytes=buf.getvalue(),
                             content_type=IMAGE_FORMAT_MIME.get(target_format.lower(), "application/octet-stream"))

    def _image_resize(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        img = Image.open(input_data)
        width, height = options.get("width"), options.get("height")
        if not width and not height:
            return EngineResult(
                ok=False, error="Provide options['width'] and/or options['height']", refundable=False,
            )

        orig_w, orig_h = img.size
        if width and not height:
            height = round(orig_h * (int(width) / orig_w))
        elif height and not width:
            width = round(orig_w * (int(height) / orig_h))

        resized = img.resize((int(width), int(height)))
        buf = io.BytesIO()
        resized.save(buf, format=img.format or "PNG")
        return EngineResult(ok=True, output_bytes=buf.getvalue(),
                             content_type=Image.MIME.get(img.format, "image/png"))

    def _image_compress(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        quality = int(options.get("quality", 60))
        img = Image.open(input_data).convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        return EngineResult(ok=True, output_bytes=buf.getvalue(), content_type="image/jpeg",
                             meta={"quality": quality})

    def _image_crop(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        img = Image.open(input_data)
        box = options.get("box")
        if box:
            left, top, right, bottom = box
        else:
            w, h = img.size
            left, top, right, bottom = w * 0.1, h * 0.1, w * 0.9, h * 0.9

        cropped = img.crop((int(left), int(top), int(right), int(bottom)))
        buf = io.BytesIO()
        cropped.save(buf, format=img.format or "PNG")
        return EngineResult(ok=True, output_bytes=buf.getvalue(),
                             content_type=Image.MIME.get(img.format, "image/png"))

    def _image_watermark(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        text = options.get("text", "TweakHub")
        img = Image.open(input_data).convert("RGBA")
        overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)

        font_path = find_ttf_font()
        font_size = max(20, img.size[0] // 15)
        font = ImageFont.truetype(font_path, font_size) if font_path else ImageFont.load_default()

        draw.text((img.size[0] * 0.1, img.size[1] * 0.85), text, fill=(255, 255, 255, 160), font=font)
        watermarked = Image.alpha_composite(img, overlay).convert("RGB")

        buf = io.BytesIO()
        watermarked.save(buf, format="JPEG", quality=90)
        return EngineResult(ok=True, output_bytes=buf.getvalue(), content_type="image/jpeg")

    def _image_rotate(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        angle = int(options.get("angle", 90))
        img = Image.open(input_data)
        rotated = img.rotate(-angle, expand=True)
        buf = io.BytesIO()
        rotated.save(buf, format=img.format or "PNG")
        return EngineResult(ok=True, output_bytes=buf.getvalue(),
                             content_type=Image.MIME.get(img.format, "image/png"))

    def _heic_to_jpg(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        import pillow_heif

        pillow_heif.register_heif_opener()
        img = Image.open(input_data).convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=90)
        return EngineResult(ok=True, output_bytes=buf.getvalue(), content_type="image/jpeg")

    def _image_bg_remove(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        """rembg (MIT) running the u2netp ONNX model (Apache-2.0 — see the
        module docstring for the license check on both). Real inference,
        not a heuristic: verified end to end against a synthetic test
        image (solid subject on a plain background) with the background
        pixels' alpha channel confirmed transparent and the subject's
        confirmed opaque, not just "didn't crash"."""
        from rembg import remove

        out_bytes = remove(input_data.read(), session=_rembg_session())
        return EngineResult(ok=True, output_bytes=out_bytes, content_type="image/png")

    def _svg_to_png(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        import cairosvg

        buf = io.BytesIO()
        cairosvg.svg2png(
            bytestring=input_data.read(), write_to=buf,
            output_width=options.get("width"), output_height=options.get("height"),
        )
        return EngineResult(ok=True, output_bytes=buf.getvalue(), content_type="image/png")

    # -- video (ffmpeg) -------------------------------------------------------

    def _run_ffmpeg(self, input_bytes: bytes, in_ext: str, out_ext: str, args: list[str],
                     extra_inputs: list[bytes] | None = None, extra_ext: str = "mp4",
                     tool_name: str | None = None) -> bytes:
        # tool_name drives a per-tool timeout (services/tool_timeouts.py)
        # instead of one blanket value for every ffmpeg call — a cheap
        # stream-copy op (video_mute) doesn't need the same ceiling as a
        # full re-encode (video_compress). Falls back to the category
        # default if a caller doesn't pass tool_name for some reason.
        timeout = get_subprocess_timeout_seconds(tool_name) if tool_name else 180
        with scratch_dir() as d:
            src = d / f"in.{in_ext}"
            src.write_bytes(input_bytes)
            cmd = ["ffmpeg", "-y", "-i", str(src)]
            for i, extra in enumerate(extra_inputs or []):
                extra_path = d / f"extra_{i}.{extra_ext}"
                extra_path.write_bytes(extra)
                cmd += ["-i", str(extra_path)]
            out = d / f"out.{out_ext}"
            cmd += [*args, str(out)]
            run(cmd, timeout=timeout)
            return out.read_bytes()

    def _video_compress(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        out = self._run_ffmpeg(input_data.read(), "mp4", "mp4",
                                ["-vcodec", "libx264", "-crf", "28", "-preset", "fast", "-acodec", "aac"],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="video/mp4")

    def _video_convert(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        target_format = options.get("target_format")
        if not target_format:
            return EngineResult(ok=False, error="options['target_format'] is required", refundable=False)
        out = self._run_ffmpeg(input_data.read(), "mp4", target_format, [], tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type=f"video/{target_format}")

    def _video_trim(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        start = str(options.get("start", "0"))
        duration = options.get("duration")
        args = ["-ss", start]
        # `if duration:` treats an explicit duration=0 the same as "not
        # given" (0 is falsy) and silently trims to the end of the file
        # instead of producing a (near-)empty clip — `is not None` is the
        # actual "was this provided" check.
        if duration is not None:
            args += ["-t", str(duration)]
        args += ["-c", "copy"]
        out = self._run_ffmpeg(input_data.read(), "mp4", "mp4", args, tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="video/mp4")

    def _video_to_gif(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        out = self._run_ffmpeg(input_data.read(), "mp4", "gif",
                                ["-vf", "fps=10,scale=480:-1:flags=lanczos"],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="image/gif")

    def _video_extract_audio(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        out = self._run_ffmpeg(input_data.read(), "mp4", "mp3", ["-vn", "-acodec", "libmp3lame"],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="audio/mpeg")

    def _video_merge(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        extra_files: list[bytes] = options.get("extra_files") or []
        if not extra_files:
            return EngineResult(
                ok=False, error="video_merge needs at least one file in options['extra_files']",
                refundable=False,
            )

        n = 1 + len(extra_files)
        filter_complex = "".join(f"[{i}:v:0][{i}:a:0]" for i in range(n)) + f"concat=n={n}:v=1:a=1[v][a]"
        out = self._run_ffmpeg(
            input_data.read(), "mp4", "mp4",
            ["-filter_complex", filter_complex, "-map", "[v]", "-map", "[a]"],
            extra_inputs=extra_files, extra_ext="mp4", tool_name=options.get("tool_name"),
        )
        return EngineResult(ok=True, output_bytes=out, content_type="video/mp4")

    def _video_resize(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        width, height = options.get("width", -2), options.get("height", -2)
        out = self._run_ffmpeg(input_data.read(), "mp4", "mp4", ["-vf", f"scale={width}:{height}"],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="video/mp4")

    def _video_watermark(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        text = options.get("text", "TweakHub").replace(":", r"\:").replace("'", r"\'")
        font_path = find_ttf_font()
        drawtext = f"drawtext=text='{text}':fontcolor=white:fontsize=24:x=10:y=h-th-10"
        if font_path:
            drawtext += f":fontfile={font_path}"
        out = self._run_ffmpeg(input_data.read(), "mp4", "mp4", ["-vf", drawtext],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="video/mp4")

    def _video_mute(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        out = self._run_ffmpeg(input_data.read(), "mp4", "mp4", ["-an", "-c:v", "copy"],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="video/mp4")

    def _subtitle_burn(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        extra_files: list[bytes] = options.get("extra_files") or []
        if not extra_files:
            return EngineResult(
                ok=False, error="subtitle_burn needs an .srt file in options['extra_files']",
                refundable=False,
            )

        timeout = get_subprocess_timeout_seconds(options.get("tool_name") or "subtitle_burn")
        with scratch_dir() as d:
            src = d / "in.mp4"
            src.write_bytes(input_data.read())
            srt = d / "subs.srt"
            srt.write_bytes(extra_files[0])
            out = d / "out.mp4"
            run(["ffmpeg", "-y", "-i", str(src), "-vf", f"subtitles={srt}", str(out)], timeout=timeout)
            return EngineResult(ok=True, output_bytes=out.read_bytes(), content_type="video/mp4")

    # -- audio (ffmpeg) -------------------------------------------------------

    def _audio_convert(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        target_format = options.get("target_format")
        if not target_format:
            return EngineResult(ok=False, error="options['target_format'] is required", refundable=False)
        out = self._run_ffmpeg(input_data.read(), "wav", target_format, [], tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type=f"audio/{target_format}")

    def _audio_compress(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        bitrate = options.get("bitrate", "96k")
        out = self._run_ffmpeg(input_data.read(), "wav", "mp3", ["-b:a", str(bitrate)],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="audio/mpeg")

    def _audio_trim(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        start = str(options.get("start", "0"))
        duration = options.get("duration")
        args = ["-ss", start]
        # See _video_trim's comment: `is not None`, not truthiness — a
        # caller-supplied duration=0 must not be reinterpreted as "no
        # duration given."
        if duration is not None:
            args += ["-t", str(duration)]
        out = self._run_ffmpeg(input_data.read(), "wav", "wav", args, tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="audio/wav")

    def _audio_merge(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        extra_files: list[bytes] = options.get("extra_files") or []
        if not extra_files:
            return EngineResult(
                ok=False, error="audio_merge needs at least one file in options['extra_files']",
                refundable=False,
            )

        n = 1 + len(extra_files)
        filter_complex = "".join(f"[{i}:a:0]" for i in range(n)) + f"concat=n={n}:v=0:a=1[a]"
        out = self._run_ffmpeg(
            input_data.read(), "wav", "wav",
            ["-filter_complex", filter_complex, "-map", "[a]"],
            extra_inputs=extra_files, extra_ext="wav", tool_name=options.get("tool_name"),
        )
        return EngineResult(ok=True, output_bytes=out, content_type="audio/wav")

    def _audio_normalize(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        out = self._run_ffmpeg(input_data.read(), "wav", "wav", ["-af", "loudnorm=I=-16:TP=-1.5:LRA=11"],
                                tool_name=options.get("tool_name"))
        return EngineResult(ok=True, output_bytes=out, content_type="audio/wav")

    # -- spreadsheets ---------------------------------------------------------

    def _csv_to_xlsx(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        from openpyxl import Workbook

        text = input_data.read().decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))

        wb = Workbook()
        ws = wb.active
        row_count = 0
        for row in reader:
            ws.append(row)
            row_count += 1

        buf = io.BytesIO()
        wb.save(buf)
        return EngineResult(
            ok=True, output_bytes=buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            meta={"rows": row_count},
        )

    def _xlsx_to_csv(self, input_data: BinaryIO, options: dict[str, Any]) -> EngineResult:
        # CSV has no concept of multiple sheets, so only the workbook's
        # active sheet is exported — every other sheet's data is dropped.
        # That's an inherent format limitation (same one document_convert.
        # py's xls_to_csv/ods_to_csv LibreOffice-filter-based handlers
        # have — see docs/engines.md), not something a different library
        # call fixes, but it was previously silent: nothing in the
        # response told the caller a multi-sheet workbook lost data.
        # `meta.sheets_total`/`meta.sheets_exported` surface it instead of
        # a `.csv` that quietly looks complete.
        from openpyxl import load_workbook

        wb = load_workbook(input_data, data_only=True)
        ws = wb.active

        buf = io.StringIO()
        writer = csv.writer(buf)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
            row_count += 1

        return EngineResult(
            ok=True, output_bytes=buf.getvalue().encode("utf-8"), content_type="text/csv",
            meta={"rows": row_count, "sheets_total": len(wb.sheetnames), "sheets_exported": 1},
        )
